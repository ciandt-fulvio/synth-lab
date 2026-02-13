"""Export simulation data to Excel (.xlsx) for Monte Carlo formula calibration.

Generates an Excel workbook with live formulas matching the actual simulation engine:
    - Sigmoid probability model: P = 1/(1+exp(-z)) where z = intercept + aw*appeals - bw*barriers
    - Non-linear perceived_risk: 1 - (1-irreversibility)(1-risk_aversion)
    - network_bonus as appeal (not barrier)
    - 7 barriers + 4 appeals = 11 emergent states
    - Gating mechanisms (trust_gate, risk_gate, value_gate) based on feature_types

Sheets:
    - Parametros: Editable sigmoid weights + gate parameters + experiment metadata
    - Mecanismos: 9 mechanism values from the experiment's scorecard
    - Simulacao: 1 row per synth with sensitivities (data) + emergent states + P(adopted) (formulas)

Usage:
    DATABASE_URL="postgresql://synthlab:synthlab@localhost:5432/synthlab" \
        uv run python scripts/export_simulation_excel.py <experiment_id>

References:
    - openpyxl docs: https://openpyxl.readthedocs.io/en/stable/
    - Emergent state formulas: src/synth_lab/services/simulation/emergent_calculator.py
    - Adoption probability: src/synth_lab/services/simulation/feature_monte_carlo.py
"""

import argparse
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

# ---------------------------------------------------------------------------
# Constants (must match feature_monte_carlo.py exactly)
# ---------------------------------------------------------------------------

# 9 mechanisms in the order used by FeatureMechanisms
MECHANISM_KEYS = [
    "irreversibility",
    "network_effect",
    "institutional_trust",
    "habit_displacement",
    "learning_curve",
    "social_visibility",
    "intrinsic_value",
    "operational_friction",
    "frequency_of_use",
]

# 9 sensitivities in the order used by UserSensitivities
SENSITIVITY_KEYS = [
    "risk_aversion",
    "social_dependency",
    "institutional_trust_level",
    "habit_plasticity",
    "friction_tolerance",
    "pragmatism",
    "digital_capability",
    "motor_ability",
    "subject_domain",
]

# 11 emergent states: (name, mechanism_key, sensitivity_key, formula_type)
# formula_type:
#   "affinity"    = mech * sens
#   "resistance"  = mech * (1 - sens)
#   "nonlinear"   = 1 - (1 - mech) * (1 - sens)  [OR gate]
EMERGENT_STATES = [
    # 7 barriers
    ("perceived_risk", "irreversibility", "risk_aversion", "nonlinear"),
    ("social_pressure", "social_visibility", "social_dependency", "affinity"),
    ("trust_barrier", "institutional_trust", "institutional_trust_level", "resistance"),
    ("habit_resistance", "habit_displacement", "habit_plasticity", "resistance"),
    ("learning_frustration", "learning_curve", "digital_capability", "resistance"),
    ("friction_burden", "operational_friction", "friction_tolerance", "resistance"),
    ("motor_barrier", "operational_friction", "motor_ability", "resistance"),
    # 4 appeals
    ("intrinsic_appeal", "intrinsic_value", "pragmatism", "affinity"),
    ("frequency_value", "frequency_of_use", "pragmatism", "affinity"),
    ("domain_advantage", "intrinsic_value", "subject_domain", "affinity"),
    ("network_bonus", "network_effect", "social_dependency", "affinity"),
]

NUM_BARRIERS = 7
NUM_APPEALS = 4

# Gate definitions (must match GATE_TYPE_MAP in feature_monte_carlo.py)
GATE_TYPE_MAP: dict[str, list[str]] = {
    "trust_gate": ["financial", "identity", "security"],
    "risk_gate": ["financial", "identity"],
    "value_gate": ["aesthetic", "flow"],
}

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
PARAM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GATE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------


def get_database_url() -> str:
    url = os.environ.get("DATABASE_URL")
    if not url:
        print("ERROR: DATABASE_URL environment variable is required")
        sys.exit(1)
    return url


def load_experiment(engine, experiment_id: str) -> dict:
    """Load experiment with scorecard_data.mechanisms and feature_types."""
    with engine.connect() as conn:
        row = conn.execute(
            text("""
                SELECT id, name, hypothesis, scorecard_data, synth_group_id
                FROM experiments
                WHERE id = :eid AND status = 'active'
            """),
            {"eid": experiment_id},
        ).fetchone()

    if not row:
        print(f"ERROR: Experiment '{experiment_id}' not found or not active")
        sys.exit(1)

    scorecard = row[3] or {}
    mechanisms = scorecard.get("mechanisms", {})
    if not mechanisms:
        print(f"ERROR: Experiment '{experiment_id}' has no mechanisms in scorecard_data")
        sys.exit(1)

    feature_types = scorecard.get("feature_types", [])

    return {
        "id": row[0],
        "name": row[1],
        "hypothesis": row[2],
        "mechanisms": mechanisms,
        "feature_types": feature_types,
        "synth_group_id": row[4],
    }


def load_synths(engine, synth_group_id: str) -> list[dict]:
    """Load all synths from a synth group with their sensitivities."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT id, nome, data
                FROM synths
                WHERE synth_group_id = :gid
                ORDER BY nome
            """),
            {"gid": synth_group_id},
        ).fetchall()

    synths = []
    for row in rows:
        data = row[2] or {}
        sensitivities = data.get("sensitivities", {})
        # Skip _meta key if present
        sens_clean = {k: v for k, v in sensitivities.items() if k != "_meta"}
        synths.append({
            "id": row[0],
            "nome": row[1],
            "sensitivities": sens_clean,
        })

    if not synths:
        print(f"ERROR: No synths found in synth_group '{synth_group_id}'")
        sys.exit(1)

    return synths


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------


def _detect_active_gates(feature_types: list[str]) -> dict[str, bool]:
    """Determine which gates are active based on feature_types."""
    type_set = set(feature_types) if feature_types else set()
    return {
        gate_name: bool(type_set & set(gate_types))
        for gate_name, gate_types in GATE_TYPE_MAP.items()
    }


def create_parametros_sheet(wb: Workbook, experiment: dict) -> None:
    """Create the 'Parametros' sheet with editable sigmoid + gate parameters.

    Layout (Parametros!B row references used in Simulacao formulas):
        B2 = intercept     (-0.2)
        B3 = barrier_weight (1.1)
        B4 = appeal_weight  (1.4)
        B6 = trust_gate_active (1 or 0)
        B7 = risk_gate_active  (1 or 0)
        B8 = value_gate_active (1 or 0)
    """
    ws = wb.active
    ws.title = "Parametros"

    # Headers
    ws["A1"] = "Parametro"
    ws["B1"] = "Valor"
    ws["C1"] = "Nota"
    ws["A1"].font = HEADER_FONT
    ws["B1"].font = HEADER_FONT
    ws["C1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["B1"].fill = HEADER_FILL
    ws["C1"].fill = HEADER_FILL

    # Sigmoid parameters (yellow background) — rows 2-4
    sigmoid_params = [
        ("intercept", -0.2, "Bias do sigmoid (z = intercept + aw*appeals - bw*barriers)"),
        ("barrier_weight", 1.1, "Peso multiplicador das barreiras no sigmoid"),
        ("appeal_weight", 1.4, "Peso multiplicador dos apelos no sigmoid"),
    ]
    for i, (name, value, note) in enumerate(sigmoid_params, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = value
        ws[f"B{i}"].fill = PARAM_FILL
        ws[f"B{i}"].number_format = "0.00"
        ws[f"C{i}"] = note
        ws[f"C{i}"].font = Font(italic=True, color="808080")

    # Blank row 5
    ws["A5"] = ""

    # Gate parameters (orange background) — rows 6-8
    active_gates = _detect_active_gates(experiment.get("feature_types", []))
    gate_params = [
        ("trust_gate_active", 1 if active_gates["trust_gate"] else 0,
         f"1=ativo para {GATE_TYPE_MAP['trust_gate']}"),
        ("risk_gate_active", 1 if active_gates["risk_gate"] else 0,
         f"1=ativo para {GATE_TYPE_MAP['risk_gate']}"),
        ("value_gate_active", 1 if active_gates["value_gate"] else 0,
         f"1=ativo para {GATE_TYPE_MAP['value_gate']}"),
    ]
    for i, (name, value, note) in enumerate(gate_params, start=6):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = value
        ws[f"B{i}"].fill = GATE_FILL
        ws[f"B{i}"].number_format = "0"
        ws[f"C{i}"] = note
        ws[f"C{i}"].font = Font(italic=True, color="808080")

    # Blank row 9
    ws["A9"] = ""

    # Experiment metadata (read-only info) — rows 10-13
    meta = [
        ("experiment_id", experiment["id"]),
        ("experiment_name", experiment["name"]),
        ("hypothesis", experiment["hypothesis"]),
        ("feature_types", ", ".join(experiment.get("feature_types", []) or ["(none)"])),
    ]
    for i, (label, value) in enumerate(meta, start=10):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"A{i}"].font = Font(italic=True, color="808080")
        ws[f"B{i}"].font = Font(italic=True, color="808080")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 55


def create_mecanismos_sheet(wb: Workbook, mechanisms: dict) -> None:
    """Create the 'Mecanismos' sheet with 9 mechanism values."""
    ws = wb.create_sheet("Mecanismos")

    ws["A1"] = "mecanismo"
    ws["B1"] = "valor"
    ws["A1"].font = HEADER_FONT
    ws["B1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["B1"].fill = HEADER_FILL

    for i, key in enumerate(MECHANISM_KEYS, start=2):
        ws[f"A{i}"] = key
        ws[f"B{i}"] = float(mechanisms.get(key, 0.0))
        ws[f"B{i}"].number_format = "0.00"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10


def create_simulacao_sheet(wb: Workbook, synths: list[dict]) -> None:
    """Create the 'Simulacao' sheet with data + Excel formulas.

    Column layout:
        A=synth_id, B=nome,
        C..K = 9 sensitivities (data),
        L..V = 11 emergent states (formulas),
        W = z_score (sigmoid input),
        X = P_base (sigmoid output, before gates),
        Y = trust_gate_multiplier,
        Z = risk_gate_multiplier,
        AA = value_gate_multiplier,
        AB = P(adopted) (final = P_base * gates)
    """
    ws = wb.create_sheet("Simulacao")

    # --- Build column layout ---
    headers: list[str] = ["synth_id", "nome"]
    headers.extend(SENSITIVITY_KEYS)
    headers.extend(s[0] for s in EMERGENT_STATES)
    headers.extend([
        "z_score", "P_base",
        "trust_gate", "risk_gate", "value_gate",
        "P(adopted)",
    ])

    # Write headers (row 1)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Color formula columns with green fill (emergent + sigmoid)
    first_formula_col = 2 + len(SENSITIVITY_KEYS) + 1  # col 12
    for col_idx in range(first_formula_col, len(headers) + 1):
        ws.cell(row=1, column=col_idx).fill = FORMULA_FILL

    # Color gate columns with orange fill
    gate_start = first_formula_col + len(EMERGENT_STATES) + 2  # z_score + P_base = +2
    for col_idx in range(gate_start, gate_start + 3):
        ws.cell(row=1, column=col_idx).fill = GATE_FILL

    # Build mappings for formula generation
    sens_col: dict[str, str] = {}
    for i, key in enumerate(SENSITIVITY_KEYS):
        sens_col[key] = get_column_letter(3 + i)

    mech_row: dict[str, int] = {}
    for i, key in enumerate(MECHANISM_KEYS):
        mech_row[key] = 2 + i

    emergent_start_col = 3 + len(SENSITIVITY_KEYS)
    emergent_col_letters: list[str] = []
    for i in range(len(EMERGENT_STATES)):
        emergent_col_letters.append(get_column_letter(emergent_start_col + i))

    # Columns after emergent states
    z_col_idx = emergent_start_col + len(EMERGENT_STATES)
    p_base_col_idx = z_col_idx + 1
    trust_gate_col_idx = p_base_col_idx + 1
    risk_gate_col_idx = trust_gate_col_idx + 1
    value_gate_col_idx = risk_gate_col_idx + 1
    p_adopted_col_idx = value_gate_col_idx + 1

    z_col = get_column_letter(z_col_idx)
    p_base_col = get_column_letter(p_base_col_idx)
    trust_gate_col = get_column_letter(trust_gate_col_idx)
    risk_gate_col = get_column_letter(risk_gate_col_idx)
    value_gate_col = get_column_letter(value_gate_col_idx)

    # Barrier and appeal column letters
    barrier_cols = emergent_col_letters[:NUM_BARRIERS]
    appeal_cols = emergent_col_letters[NUM_BARRIERS:]

    # --- Write data rows ---
    for row_idx, synth in enumerate(synths, start=2):
        # A: synth_id
        ws.cell(row=row_idx, column=1, value=synth["id"])
        # B: nome
        ws.cell(row=row_idx, column=2, value=synth["nome"])

        # C-K: 9 sensitivities (data values)
        for i, key in enumerate(SENSITIVITY_KEYS):
            val = synth["sensitivities"].get(key, 0.5)
            cell = ws.cell(row=row_idx, column=3 + i, value=float(val))
            cell.number_format = "0.0000"

        # L-V: 11 emergent states (Excel formulas)
        for i, (_, mech_key, sens_key, formula_type) in enumerate(EMERGENT_STATES):
            col = emergent_start_col + i
            m_row = mech_row[mech_key]
            s_col = sens_col[sens_key]

            if formula_type == "nonlinear":
                # OR gate: 1 - (1 - mech) * (1 - sens)
                formula = f"=1-(1-Mecanismos!B{m_row})*(1-{s_col}{row_idx})"
            elif formula_type == "resistance":
                # mech * (1 - sensitivity)
                formula = f"=Mecanismos!B{m_row}*(1-{s_col}{row_idx})"
            else:
                # affinity: mech * sensitivity
                formula = f"=Mecanismos!B{m_row}*{s_col}{row_idx}"

            cell = ws.cell(row=row_idx, column=col, value=formula)
            cell.number_format = "0.0000"
            cell.fill = FORMULA_FILL

        # z_score = intercept + appeal_weight * SUM(appeals) - barrier_weight * SUM(barriers)
        barriers_sum = "+".join(f"{c}{row_idx}" for c in barrier_cols)
        appeals_sum = "+".join(f"{c}{row_idx}" for c in appeal_cols)
        z_formula = (
            f"=Parametros!B2"
            f"+Parametros!B4*({appeals_sum})"
            f"-Parametros!B3*({barriers_sum})"
        )
        cell = ws.cell(row=row_idx, column=z_col_idx, value=z_formula)
        cell.number_format = "0.0000"
        cell.fill = FORMULA_FILL

        # P_base = sigmoid(z) = 1/(1+EXP(-z))
        p_base_formula = f"=1/(1+EXP(-{z_col}{row_idx}))"
        cell = ws.cell(row=row_idx, column=p_base_col_idx, value=p_base_formula)
        cell.number_format = "0.0%"
        cell.fill = FORMULA_FILL

        # Trust gate: IF active, (0.15 + 0.85 * sigmoid(8*(trust_level - 0.45)))
        # trust_level = 1 - trust_barrier
        # trust_barrier is the 3rd emergent state (index 2)
        trust_barrier_col = emergent_col_letters[2]  # trust_barrier
        trust_gate_formula = (
            f"=IF(Parametros!B6=1,"
            f"0.15+0.85*(1/(1+EXP(-8*((1-{trust_barrier_col}{row_idx})-0.45)))),"
            f"1)"
        )
        cell = ws.cell(row=row_idx, column=trust_gate_col_idx, value=trust_gate_formula)
        cell.number_format = "0.0000"
        cell.fill = GATE_FILL

        # Risk gate: IF active, MAX(0.05, 1 - perceived_risk^1.6)
        # perceived_risk is the 1st emergent state (index 0)
        perceived_risk_col = emergent_col_letters[0]
        risk_gate_formula = (
            f"=IF(Parametros!B7=1,"
            f"MAX(0.05,1-{perceived_risk_col}{row_idx}^1.6),"
            f"1)"
        )
        cell = ws.cell(row=row_idx, column=risk_gate_col_idx, value=risk_gate_formula)
        cell.number_format = "0.0000"
        cell.fill = GATE_FILL

        # Value gate: IF active, (0.2 + 0.8 * sigmoid(9*(intrinsic_appeal - 0.25)))
        # intrinsic_appeal is the 8th emergent state (index 7)
        intrinsic_appeal_col = emergent_col_letters[7]  # intrinsic_appeal
        value_gate_formula = (
            f"=IF(Parametros!B8=1,"
            f"0.2+0.8*(1/(1+EXP(-9*({intrinsic_appeal_col}{row_idx}-0.25)))),"
            f"1)"
        )
        cell = ws.cell(row=row_idx, column=value_gate_col_idx, value=value_gate_formula)
        cell.number_format = "0.0000"
        cell.fill = GATE_FILL

        # P(adopted) = P_base * trust_gate * risk_gate * value_gate
        p_adopted_formula = (
            f"={p_base_col}{row_idx}"
            f"*{trust_gate_col}{row_idx}"
            f"*{risk_gate_col}{row_idx}"
            f"*{value_gate_col}{row_idx}"
        )
        cell = ws.cell(row=row_idx, column=p_adopted_col_idx, value=p_adopted_formula)
        cell.number_format = "0.0%"
        cell.fill = FORMULA_FILL

    # --- Column widths ---
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Freeze panes: freeze header row + first 2 columns (id, nome)
    ws.freeze_panes = "C2"


def export_excel(experiment_id: str) -> Path:
    """Main export function. Returns path to generated .xlsx file."""
    engine = create_engine(get_database_url())

    print(f"Loading experiment {experiment_id}...")
    experiment = load_experiment(engine, experiment_id)
    print(f"  Name: {experiment['name']}")
    print(f"  Mechanisms: {len(experiment['mechanisms'])} keys")
    print(f"  Feature types: {experiment['feature_types'] or '(none)'}")

    # Determine active gates
    active_gates = _detect_active_gates(experiment["feature_types"])
    active = [g for g, v in active_gates.items() if v]
    print(f"  Active gates: {active or '(none)'}")

    print(f"Loading synths from group {experiment['synth_group_id']}...")
    synths = load_synths(engine, experiment["synth_group_id"])
    print(f"  Found {len(synths)} synths")

    # Check how many have sensitivities
    with_sens = sum(1 for s in synths if s["sensitivities"])
    print(f"  With sensitivities: {with_sens}/{len(synths)}")

    # Create workbook
    wb = Workbook()
    create_parametros_sheet(wb, experiment)
    create_mecanismos_sheet(wb, experiment["mechanisms"])
    create_simulacao_sheet(wb, synths)

    # Save to exports/
    exports_dir = Path("exports")
    exports_dir.mkdir(exist_ok=True)

    # Sanitize filename
    safe_name = experiment["name"].replace("/", "_").replace("\\", "_")[:80]
    output_path = exports_dir / f"{safe_name}.xlsx"
    wb.save(str(output_path))

    print(f"\nExported to: {output_path}")
    print(f"  Sheets: Parametros, Mecanismos, Simulacao")
    print(f"  Rows in Simulacao: {len(synths)} synths")
    print(f"\nFormula: P = sigmoid(intercept + aw*appeals - bw*barriers) * gates")
    print(f"Edit Parametros!B2:B4 for sigmoid weights, B6:B8 for gate toggles.")

    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Export simulation data to Excel for Monte Carlo formula calibration"
    )
    parser.add_argument("experiment_id", help="Experiment ID (e.g., exp_5b094f89)")
    args = parser.parse_args()

    export_excel(args.experiment_id)
