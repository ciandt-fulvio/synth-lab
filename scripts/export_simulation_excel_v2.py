"""Export simulation data to Excel (.xlsx) — v2 simplified formula.

Simplified adoption model for calibration:
    B = SUM(7 barriers)
    A = SUM(4 appeals)
    z = intercept + aw*A - bw*B
    p = sigmoid(z)
    trust_gate = sigmoid(8 * (institutional_trust_level - 0.45))
    P(adopted) = p * (0.6 + 0.4 * trust_gate)

The trust gate always applies (no feature_type gating).

Sheets:
    - Parametros: Editable sigmoid weights + trust gate params
    - Mecanismos: 9 mechanism values from the experiment's scorecard
    - Simulacao: 1 row per synth with sensitivities + emergent states + P(adopted)

Usage:
    DATABASE_URL="postgresql://synthlab:synthlab@localhost:5432/synthlab" \
        uv run python scripts/export_simulation_excel_v2.py <experiment_id>

References:
    - openpyxl docs: https://openpyxl.readthedocs.io/en/stable/
    - Emergent state formulas: src/synth_lab/services/simulation/emergent_calculator.py
    - Adoption probability: src/synth_lab/services/simulation/feature_monte_carlo.py
"""

import argparse
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MECHANISM_KEYS = [
    "irreversibility",
    "network_effect",
    "institutional_trust",
    "habit_displacement",
    "learning_curve",
    "social_visibility",
    "intrinsic_value",
    "operational_friction",
    "frequency_of_use",
]

SENSITIVITY_KEYS = [
    "risk_aversion",
    "social_dependency",
    "institutional_trust_level",
    "habit_plasticity",
    "friction_tolerance",
    "pragmatism",
    "digital_capability",
    "motor_ability",
    "subject_domain",
]

# 11 emergent states: (name, mechanism_key, sensitivity_key, formula_type)
# "nonlinear" = 1 - (1-mech)(1-sens)   [OR gate]
# "resistance" = mech * (1-sens)
# "affinity"   = mech * sens
EMERGENT_STATES = [
    # 7 barriers
    ("perceived_risk", "irreversibility", "risk_aversion", "nonlinear"),
    ("social_pressure", "social_visibility", "social_dependency", "affinity"),
    ("trust_barrier", "institutional_trust", "institutional_trust_level", "resistance"),
    ("habit_resistance", "habit_displacement", "habit_plasticity", "resistance"),
    ("learning_frustration", "learning_curve", "digital_capability", "resistance"),
    ("friction_burden", "operational_friction", "friction_tolerance", "resistance"),
    ("motor_barrier", "operational_friction", "motor_ability", "resistance"),
    # 4 appeals
    ("intrinsic_appeal", "intrinsic_value", "pragmatism", "affinity"),
    ("frequency_value", "frequency_of_use", "pragmatism", "affinity"),
    ("domain_advantage", "intrinsic_value", "subject_domain", "affinity"),
    ("network_bonus", "network_effect", "social_dependency", "affinity"),
]

NUM_BARRIERS = 7
NUM_APPEALS = 4

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
PARAM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GATE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------


def get_database_url() -> str:
    url = os.environ.get("DATABASE_URL")
    if not url:
        print("ERROR: DATABASE_URL environment variable is required")
        sys.exit(1)
    return url


def load_experiment(engine, experiment_id: str) -> dict:
    """Load experiment with scorecard_data.mechanisms."""
    with engine.connect() as conn:
        row = conn.execute(
            text("""
                SELECT id, name, hypothesis, scorecard_data, synth_group_id
                FROM experiments
                WHERE id = :eid AND status = 'active'
            """),
            {"eid": experiment_id},
        ).fetchone()

    if not row:
        print(f"ERROR: Experiment '{experiment_id}' not found or not active")
        sys.exit(1)

    scorecard = row[3] or {}
    mechanisms = scorecard.get("mechanisms", {})
    if not mechanisms:
        print(f"ERROR: Experiment '{experiment_id}' has no mechanisms in scorecard_data")
        sys.exit(1)

    return {
        "id": row[0],
        "name": row[1],
        "hypothesis": row[2],
        "mechanisms": mechanisms,
        "synth_group_id": row[4],
    }


def load_synths(engine, synth_group_id: str) -> list[dict]:
    """Load all synths from a synth group with their sensitivities."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT id, nome, data
                FROM synths
                WHERE synth_group_id = :gid
                ORDER BY nome
            """),
            {"gid": synth_group_id},
        ).fetchall()

    synths = []
    for row in rows:
        data = row[2] or {}
        sensitivities = data.get("sensitivities", {})
        sens_clean = {k: v for k, v in sensitivities.items() if k != "_meta"}
        synths.append({
            "id": row[0],
            "nome": row[1],
            "sensitivities": sens_clean,
        })

    if not synths:
        print(f"ERROR: No synths found in synth_group '{synth_group_id}'")
        sys.exit(1)

    return synths


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------


def create_parametros_sheet(wb: Workbook, experiment: dict) -> None:
    """Create the 'Parametros' sheet.

    Layout (row references used by Simulacao formulas):
        B2 = intercept       (-0.2)
        B3 = barrier_weight  (1.1)
        B4 = appeal_weight   (1.4)
        B6 = trust_gate_steepness (8)
        B7 = trust_gate_midpoint  (0.45)
        B8 = trust_gate_floor     (0.6)
        B9 = trust_gate_range     (0.4)
    """
    ws = wb.active
    ws.title = "Parametros"

    ws["A1"] = "Parametro"
    ws["B1"] = "Valor"
    ws["C1"] = "Nota"
    for c in ("A1", "B1", "C1"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL

    # Sigmoid parameters — rows 2-4
    sigmoid_params = [
        ("intercept", -0.2, "z = intercept + aw*A - bw*B"),
        ("barrier_weight", 1.1, "bw: peso das barreiras"),
        ("appeal_weight", 1.4, "aw: peso dos apelos"),
    ]
    for i, (name, value, note) in enumerate(sigmoid_params, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = value
        ws[f"B{i}"].fill = PARAM_FILL
        ws[f"B{i}"].number_format = "0.00"
        ws[f"C{i}"] = note
        ws[f"C{i}"].font = Font(italic=True, color="808080")

    # Trust gate parameters — rows 6-9
    gate_params = [
        ("trust_gate_steepness", 8, "Inclinação do sigmoid do gate"),
        ("trust_gate_midpoint", 0.45, "Ponto de inflexão (trust_level)"),
        ("trust_gate_floor", 0.6, "Multiplicador mínimo (floor)"),
        ("trust_gate_range", 0.4, "Faixa acima do floor (floor+range=1.0)"),
    ]
    for i, (name, value, note) in enumerate(gate_params, start=6):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = value
        ws[f"B{i}"].fill = GATE_FILL
        ws[f"B{i}"].number_format = "0.00"
        ws[f"C{i}"] = note
        ws[f"C{i}"].font = Font(italic=True, color="808080")

    # Experiment metadata — rows 11-13
    meta = [
        ("experiment_id", experiment["id"]),
        ("experiment_name", experiment["name"]),
        ("hypothesis", experiment["hypothesis"]),
    ]
    for i, (label, value) in enumerate(meta, start=11):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"A{i}"].font = Font(italic=True, color="808080")
        ws[f"B{i}"].font = Font(italic=True, color="808080")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50


def create_mecanismos_sheet(wb: Workbook, mechanisms: dict) -> None:
    """Create the 'Mecanismos' sheet with 9 mechanism values."""
    ws = wb.create_sheet("Mecanismos")

    ws["A1"] = "mecanismo"
    ws["B1"] = "valor"
    ws["A1"].font = HEADER_FONT
    ws["B1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["B1"].fill = HEADER_FILL

    for i, key in enumerate(MECHANISM_KEYS, start=2):
        ws[f"A{i}"] = key
        ws[f"B{i}"] = float(mechanisms.get(key, 0.0))
        ws[f"B{i}"].number_format = "0.00"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10


def create_simulacao_sheet(wb: Workbook, synths: list[dict]) -> None:
    """Create the 'Simulacao' sheet.

    Columns:
        A=synth_id, B=nome,
        C..K = 9 sensitivities (data),
        L..V = 11 emergent states (formulas),
        W = z_score,
        X = P_base (sigmoid),
        Y = trust_gate,
        Z = P(adopted) = P_base * (floor + range * trust_gate)
    """
    ws = wb.create_sheet("Simulacao")

    headers: list[str] = ["synth_id", "nome"]
    headers.extend(SENSITIVITY_KEYS)
    headers.extend(s[0] for s in EMERGENT_STATES)
    headers.extend(["z_score", "P_base", "trust_gate", "P(adopted)"])

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Green fill for formula columns (emergent states onward)
    first_formula_col = 2 + len(SENSITIVITY_KEYS) + 1
    for col_idx in range(first_formula_col, len(headers) + 1):
        ws.cell(row=1, column=col_idx).fill = FORMULA_FILL

    # Orange fill for trust_gate header
    trust_gate_header_col = first_formula_col + len(EMERGENT_STATES) + 2  # z + P_base = +2
    ws.cell(row=1, column=trust_gate_header_col).fill = GATE_FILL

    # Mappings
    sens_col: dict[str, str] = {}
    for i, key in enumerate(SENSITIVITY_KEYS):
        sens_col[key] = get_column_letter(3 + i)

    mech_row: dict[str, int] = {}
    for i, key in enumerate(MECHANISM_KEYS):
        mech_row[key] = 2 + i

    emergent_start_col = 3 + len(SENSITIVITY_KEYS)
    emergent_col_letters: list[str] = []
    for i in range(len(EMERGENT_STATES)):
        emergent_col_letters.append(get_column_letter(emergent_start_col + i))

    barrier_cols = emergent_col_letters[:NUM_BARRIERS]
    appeal_cols = emergent_col_letters[NUM_BARRIERS:]

    z_col_idx = emergent_start_col + len(EMERGENT_STATES)
    p_base_col_idx = z_col_idx + 1
    trust_gate_col_idx = p_base_col_idx + 1
    p_adopted_col_idx = trust_gate_col_idx + 1

    z_col = get_column_letter(z_col_idx)
    p_base_col = get_column_letter(p_base_col_idx)
    trust_gate_col = get_column_letter(trust_gate_col_idx)

    # institutional_trust_level is the 3rd sensitivity (index 2 → column E)
    itl_col = sens_col["institutional_trust_level"]

    for row_idx, synth in enumerate(synths, start=2):
        ws.cell(row=row_idx, column=1, value=synth["id"])
        ws.cell(row=row_idx, column=2, value=synth["nome"])

        # Sensitivities (data)
        for i, key in enumerate(SENSITIVITY_KEYS):
            val = synth["sensitivities"].get(key, 0.5)
            cell = ws.cell(row=row_idx, column=3 + i, value=float(val))
            cell.number_format = "0.0000"

        # 11 emergent states (formulas)
        for i, (_, mech_key, sens_key, formula_type) in enumerate(EMERGENT_STATES):
            col = emergent_start_col + i
            m_row = mech_row[mech_key]
            s_col = sens_col[sens_key]

            if formula_type == "nonlinear":
                formula = f"=1-(1-Mecanismos!B{m_row})*(1-{s_col}{row_idx})"
            elif formula_type == "resistance":
                formula = f"=Mecanismos!B{m_row}*(1-{s_col}{row_idx})"
            else:
                formula = f"=Mecanismos!B{m_row}*{s_col}{row_idx}"

            cell = ws.cell(row=row_idx, column=col, value=formula)
            cell.number_format = "0.0000"
            cell.fill = FORMULA_FILL

        # z = intercept + aw*A - bw*B
        barriers_sum = "+".join(f"{c}{row_idx}" for c in barrier_cols)
        appeals_sum = "+".join(f"{c}{row_idx}" for c in appeal_cols)
        z_formula = (
            f"=Parametros!B2"
            f"+Parametros!B4*({appeals_sum})"
            f"-Parametros!B3*({barriers_sum})"
        )
        cell = ws.cell(row=row_idx, column=z_col_idx, value=z_formula)
        cell.number_format = "0.0000"
        cell.fill = FORMULA_FILL

        # P_base = sigmoid(z)
        cell = ws.cell(
            row=row_idx, column=p_base_col_idx,
            value=f"=1/(1+EXP(-{z_col}{row_idx}))",
        )
        cell.number_format = "0.0%"
        cell.fill = FORMULA_FILL

        # trust_gate = sigmoid(steepness * (institutional_trust_level - midpoint))
        # Uses the raw sensitivity value directly (not the emergent trust_barrier)
        # Parametros!B6=steepness, B7=midpoint
        cell = ws.cell(
            row=row_idx, column=trust_gate_col_idx,
            value=f"=1/(1+EXP(-Parametros!B6*({itl_col}{row_idx}-Parametros!B7)))",
        )
        cell.number_format = "0.0000"
        cell.fill = GATE_FILL

        # P(adopted) = P_base * (floor + range * trust_gate)
        # Parametros!B8=floor, B9=range
        cell = ws.cell(
            row=row_idx, column=p_adopted_col_idx,
            value=(
                f"={p_base_col}{row_idx}"
                f"*(Parametros!B8+Parametros!B9*{trust_gate_col}{row_idx})"
            ),
        )
        cell.number_format = "0.0%"
        cell.fill = FORMULA_FILL

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = "C2"


def export_excel(experiment_id: str) -> Path:
    """Main export function."""
    engine = create_engine(get_database_url())

    print(f"Loading experiment {experiment_id}...")
    experiment = load_experiment(engine, experiment_id)
    print(f"  Name: {experiment['name']}")
    print(f"  Mechanisms: {len(experiment['mechanisms'])} keys")

    print(f"Loading synths from group {experiment['synth_group_id']}...")
    synths = load_synths(engine, experiment["synth_group_id"])
    print(f"  Found {len(synths)} synths")

    with_sens = sum(1 for s in synths if s["sensitivities"])
    print(f"  With sensitivities: {with_sens}/{len(synths)}")

    wb = Workbook()
    create_parametros_sheet(wb, experiment)
    create_mecanismos_sheet(wb, experiment["mechanisms"])
    create_simulacao_sheet(wb, synths)

    exports_dir = Path("exports")
    exports_dir.mkdir(exist_ok=True)

    safe_name = experiment["name"].replace("/", "_").replace("\\", "_")[:80]
    output_path = exports_dir / f"{safe_name} v2.xlsx"
    wb.save(str(output_path))

    print(f"\nExported to: {output_path}")
    print(f"  Sheets: Parametros, Mecanismos, Simulacao")
    print(f"  Rows: {len(synths)} synths")
    print(f"\nFormula v2:")
    print(f"  z = intercept + aw*A - bw*B")
    print(f"  p = sigmoid(z)")
    print(f"  trust_gate = sigmoid(steepness * (institutional_trust_level - midpoint))")
    print(f"  P(adopted) = p * (floor + range * trust_gate)")

    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Export simulation v2: sigmoid + trust gate only"
    )
    parser.add_argument("experiment_id", help="Experiment ID (e.g., exp_5b094f89)")
    args = parser.parse_args()

    export_excel(args.experiment_id)
